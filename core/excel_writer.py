"""Excel 출력 모듈 — 원본 표 형식 재현, 문서 타입별 스타일 자동 적용."""

import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from .postprocess import detect_cell_type, parse_numeric

# ── 스타일 프로필 ──

STYLE_PROFILES = {
    "default": {
        "header_fill": "D9E2F3",
        "header_font_color": "000000",
        "header_font_size": 11,
        "subtotal_fill": "F2F2F2",
        "data_font_size": 10,
        "number_format": "#,##0",
        "header_number_format": "#,##0",
    },
    "financial_statement": {
        "header_fill": "D9E2F3",
        "header_font_color": "000000",
        "header_font_size": 11,
        "subtotal_fill": "F2F2F2",
        "data_font_size": 10,
        "number_format": "#,##0",
        "header_number_format": "#,##0",
    },
    "demand_forecast": {
        "header_fill": "002060",
        "header_font_color": "FFFFFF",
        "header_font_size": 10,
        "subtotal_fill": "E2EFDA",
        "data_font_size": 10,
        # 원본 수요예측표 포맷: 양수=콤마, 음수=빨간, 0=대시
        "number_format": '#,##0;[Red]\\-#,##0;\\-;',
        # 헤더(스프레드): +/-부호 표시
        "header_number_format": '\\+#,##0;\\-#,##0',
    },
}


def _get_style(profile_name: str = "default") -> dict:
    return STYLE_PROFILES.get(profile_name, STYLE_PROFILES["default"])


def _make_styles(profile_name: str = "default"):
    s = _get_style(profile_name)
    return {
        "header_fill": PatternFill("solid", fgColor=s["header_fill"]),
        "header_font": Font(bold=True, size=s["header_font_size"], name="맑은 고딕",
                            color=s["header_font_color"]),
        "subtotal_fill": PatternFill("solid", fgColor=s["subtotal_fill"]),
        "subtotal_font": Font(bold=True, size=s["data_font_size"], name="맑은 고딕"),
        "category_font": Font(bold=True, size=s["data_font_size"], name="맑은 고딕"),
        "data_font": Font(size=s["data_font_size"], name="맑은 고딕"),
        "yellow_fill": PatternFill("solid", fgColor="FFFF00"),
        "title_font": Font(bold=True, size=14, name="맑은 고딕"),
        "meta_font": Font(size=10, name="맑은 고딕"),
        "meta_label_font": Font(bold=True, size=10, name="맑은 고딕"),
    }


# 테두리
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="medium", color="666666"),
    bottom=Side(style="medium", color="666666"),
)

# 행 타입 감지 패턴
SUBTOTAL_KEYWORDS = {"총계", "소계", "합계", "누적합계"}
CATEGORY_PATTERN = re.compile(r"^\[.+\]$")
DETAIL_PATTERN = re.compile(r"^[ㆍ·\-•]")
SUMMARY_KEYWORDS = {"비중(%)", "비중", "누적비중(%)", "누적비중"}


def _detect_row_type(first_cell: str) -> str:
    v = first_cell.strip()
    if not v:
        return "normal"
    if v in ("구 분", "구분"):
        return "header"
    if v in SUMMARY_KEYWORDS:
        return "summary"
    if any(kw in v for kw in SUBTOTAL_KEYWORDS):
        return "subtotal"
    if CATEGORY_PATTERN.match(v):
        return "category"
    if DETAIL_PATTERN.match(v):
        return "detail"
    if "년" in v and ("월" in v or "기" in v):
        return "period"
    return "normal"


def _cjk_width(text: str) -> int:
    width = 0
    for ch in str(text):
        if ord(ch) > 0x7F:
            width += 2
        else:
            width += 1
    return width


def _auto_column_widths(ws, style_profile: str = "default") -> None:
    """열 너비를 자동 조정한다."""
    is_demand = style_profile == "demand_forecast"

    for col_cells in ws.columns:
        max_width = 0
        col_idx = col_cells[0].column
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            cell_width = _cjk_width(cell.value) if cell.value else 0
            max_width = max(max_width, cell_width)

        if is_demand:
            # 수요예측표: 텍스트 열은 넓게, 숫자(스프레드) 열은 좁게
            if col_idx <= 6:
                ws.column_dimensions[col_letter].width = max(max_width + 3, 14)
            else:
                ws.column_dimensions[col_letter].width = max(max_width + 2, 7)
        else:
            if col_idx == 1:
                ws.column_dimensions[col_letter].width = max(max_width + 4, 30)
            else:
                ws.column_dimensions[col_letter].width = max(max_width + 3, 18)


def _apply_number_format(cell, value: str, num_fmt: str = "#,##0") -> None:
    """셀 값의 타입에 따라 숫자 포맷을 적용한다."""
    cell_type = detect_cell_type(value)
    parsed = parse_numeric(value)

    if parsed is not None and not (isinstance(parsed, float) and (parsed != parsed)):
        if cell_type == "percent":
            cell.value = parsed / 100
            cell.number_format = '0.0%'
        elif "." in value:
            cell.value = parsed
            cell.number_format = '#,##0.00'
        else:
            try:
                cell.value = int(parsed)
            except (ValueError, OverflowError):
                cell.value = parsed
            cell.number_format = num_fmt
    else:
        cell.value = value


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """DataFrame을 정리한다: Col_N 헤더 치환, 중복/쓰레기 행 제거."""
    num_cols = len(df.columns)

    has_placeholder = any(str(c).startswith("Col_") for c in df.columns)

    if has_placeholder:
        header_idx = None
        for i in range(min(len(df), 10)):
            row_vals = [str(df.iloc[i, c]).strip() for c in range(num_cols)]
            has_gubun = any(v in ("구분", "구 분", "항목") for v in row_vals)
            has_period = any(re.match(r"제\d+기", v) for v in row_vals)
            if has_gubun or has_period:
                header_idx = i
                break

        if header_idx is not None:
            new_header = [
                str(df.iloc[header_idx, c]).strip() or f"열{c+1}"
                for c in range(num_cols)
            ]
            title_df = df.iloc[:header_idx].copy()
            data_df = df.iloc[header_idx + 1:].copy()

            title_df.columns = new_header
            data_df.columns = new_header
            df = pd.concat([title_df, data_df], ignore_index=True)

    # 중복/쓰레기 행 제거
    drop_indices = []
    header_vals = set(str(c) for c in df.columns)

    for i in range(len(df)):
        row_vals = [str(df.iloc[i, c]).strip() for c in range(num_cols)]

        if any(re.match(r"Col_\d+", v) for v in row_vals):
            drop_indices.append(i)
            continue

        if set(row_vals) == header_vals:
            drop_indices.append(i)

    if drop_indices:
        df = df.drop(df.index[drop_indices]).reset_index(drop=True)

    return df


def _write_metadata(ws, metadata: dict, styles: dict) -> int:
    """메타데이터(제목, 발행정보, 시장정보)를 시트 상단에 쓴다.

    Returns:
        데이터 시작 행 번호 (1-indexed)
    """
    if not metadata:
        return 1

    row = 1

    # 제목
    title = metadata.get("title", "")
    if title:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = styles["title_font"]
        cell.alignment = Alignment(vertical="center")
        row += 1

    # 발행정보 + 시장정보 나란히
    info = metadata.get("info", [])
    market_info = metadata.get("market_info", [])
    max_rows = max(len(info), len(market_info))

    for i in range(max_rows):
        # 왼쪽: 발행정보
        if i < len(info):
            label = info[i][0] if len(info[i]) > 0 else ""
            value = info[i][1] if len(info[i]) > 1 else ""
            if label:
                ws.cell(row=row, column=1, value=f"· {label}").font = styles["meta_font"]
                ws.cell(row=row, column=3, value=str(value)).font = styles["meta_font"]

        # 오른쪽: 시장정보
        if i < len(market_info):
            label = market_info[i][0] if len(market_info[i]) > 0 else ""
            value = market_info[i][1] if len(market_info[i]) > 1 else ""
            if label:
                ws.cell(row=row, column=5, value=f"· {label}").font = styles["meta_font"]
                cell = ws.cell(row=row, column=7, value=str(value))
                cell.font = styles["meta_font"]
                # 숫자면 우측 정렬
                try:
                    cell.value = float(value)
                    cell.number_format = '0.000'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                except (ValueError, TypeError):
                    pass

        row += 1

    # 빈 행 하나
    row += 1
    return row


def write_excel(
    tables: list[pd.DataFrame],
    error_cells_per_table: list[list[tuple[int, int]]],
    output_path: str,
    style_profile: str = "default",
    metadata: dict | None = None,
) -> str:
    """추출된 표들을 원본 형식을 재현하여 Excel로 저장한다."""
    wb = Workbook()
    wb.remove(wb.active)

    tables = [_clean_dataframe(df) for df in tables]

    styles = _make_styles(style_profile)
    profile = _get_style(style_profile)
    num_fmt = profile["number_format"]
    header_num_fmt = profile["header_number_format"]

    for t_idx, df in enumerate(tables):
        sheet_name = f"표{t_idx + 1}"
        ws = wb.create_sheet(title=sheet_name)
        error_set = (
            set(error_cells_per_table[t_idx])
            if t_idx < len(error_cells_per_table)
            else set()
        )
        num_cols = len(df.columns)

        # ── 메타데이터 영역 (첫 번째 표만) ──
        data_start_row = 1
        if t_idx == 0 and metadata:
            data_start_row = _write_metadata(ws, metadata, styles)

        # ── 헤더 행 ──
        header_row = data_start_row
        for c_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=header_row, column=c_idx)

            # 스프레드 헤더(+30, -100 등) → 숫자로 표시
            try:
                num_val = int(str(col_name).replace("+", ""))
                cell.value = num_val
                cell.number_format = header_num_fmt
            except (ValueError, TypeError):
                cell.value = str(col_name)

            cell.fill = styles["header_fill"]
            cell.font = styles["header_font"]
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)

        # ── 데이터 행 ──
        for r_idx in range(len(df)):
            excel_row = header_row + 1 + r_idx

            first_val = str(df.iloc[r_idx, 0]).strip() if pd.notna(df.iloc[r_idx, 0]) else ""
            row_type = _detect_row_type(first_val)

            for c_idx in range(num_cols):
                cell = ws.cell(row=excel_row, column=c_idx + 1)
                val = str(df.iloc[r_idx, c_idx]).strip()

                # 값 설정
                if c_idx == 0:
                    cell.value = val
                else:
                    _apply_number_format(cell, val, num_fmt)

                # 테두리
                cell.border = THIN_BORDER

                # ── 행 타입별 스타일 ──
                if row_type == "header":
                    cell.fill = styles["header_fill"]
                    cell.font = styles["header_font"]
                    cell.border = HEADER_BORDER
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                elif row_type in ("subtotal", "summary"):
                    if row_type == "subtotal":
                        cell.fill = styles["subtotal_fill"]
                    cell.font = styles["subtotal_font"]
                    cell.border = Border(
                        left=Side(style="thin", color="CCCCCC"),
                        right=Side(style="thin", color="CCCCCC"),
                        top=Side(style="thin", color="999999"),
                        bottom=Side(style="thin", color="999999"),
                    )
                    cell.alignment = Alignment(
                        horizontal="center" if c_idx == 0 else "right",
                        vertical="center",
                    )

                elif row_type == "category":
                    cell.font = styles["category_font"]
                    cell.alignment = Alignment(
                        horizontal="left" if c_idx == 0 else "right",
                        vertical="center",
                    )

                elif row_type == "detail":
                    cell.font = styles["data_font"]
                    if c_idx == 0:
                        cell.alignment = Alignment(indent=2, vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")

                elif row_type == "period":
                    cell.font = Font(size=styles["data_font"].size, name="맑은 고딕",
                                     italic=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                else:
                    cell.font = styles["data_font"]
                    if c_idx == 0:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # 오류 셀 하이라이트
                if (r_idx, c_idx) in error_set:
                    cell.fill = styles["yellow_fill"]

        # ── 마무리 ──
        _auto_column_widths(ws, style_profile)
        ws.freeze_panes = f"A{header_row + 1}"
        ws.sheet_properties.defaultRowHeight = 20

    if not tables:
        ws = wb.create_sheet(title="결과")
        ws.cell(row=1, column=1, value="추출된 표가 없습니다.")

    wb.save(output_path)
    return output_path
